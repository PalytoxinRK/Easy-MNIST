import tensorflow as tf
#图片处理库
from PIL import Image
import numpy as np
import os
#标签处理库
from openpyxl import load_workbook
# import scipy


#图片数据 函数参数表示图片的起始和终结，比如传入1,4000就是读取第一张到第4000张图片
def loadImageAndToMatrix(first, second):
	print("START IMAGE")
	dir = './Data/'
	filename = '.jpg'
	#读入第一张图片
	path = dir + str(first) + filename
	im = Image.open(path)
	im = im.convert("1")
	#矩阵形式储存
	matrix = np.matrix(im.getdata())

	for i in range(first+1,second+1):
		path = dir + str(i) + filename
		#print(path)
		im = Image.open(path)
		im = im.convert("1") 
		data = im.getdata()
		data = np.matrix(data)
		matrix = np.row_stack((matrix,data))
	
	print(matrix.shape)
	
	return matrix
	
#标签数据 函数参数表示标签值的起始和终结，比如传入1,4000就是读取第一张到第4000张图片的标签值
def loadYToMatrix(first, second):
	print("START Y")
	book = load_workbook(filename="./Label.xlsx")
	 #读取名字为Sheet1的表
	sheet = book.get_sheet_by_name("Sheet1")
	#用于存储数据的数组
	#读取第一列数据
	Label = []
	row_num = 1
	while row_num <= 10 :
		Label.append(sheet.cell(row=row_num, column=first).value)
		row_num = row_num + 1
	Label = np.mat(Label)
	#print(Label.shape)
	for col_num in range(first+1,second+1):
		row_num = 1
		data = []
		while row_num <= 10 :
			data.append(sheet.cell(row=row_num, column=col_num).value)
			row_num = row_num + 1
		#print(data)
		data = np.mat(data)
		Label = np.row_stack((Label, data))
	print(Label.shape)
	
	return Label

#构建卷积网络模型
class CNN:
	
	def __init__(self):
	
		self.createNetwork()
		# 保存和加载网络模型
		self.saver = tf.train.Saver()
		self.sess = tf.InteractiveSession()
		self.sess.run(tf.initialize_all_variables())
		#如果检查点存在就载入已经有的模型
		checkpoint = tf.train.get_checkpoint_state("Model")
		if checkpoint and checkpoint.model_checkpoint_path:
			self.saver.restore(self.sess, checkpoint.model_checkpoint_path)
			print("Successfully loaded:", checkpoint.model_checkpoint_path)
		else:
			print("Could not find old network weights")
	
	#构建CNN卷积神经网络
	def weight_variable(self, shape):
		initial = tf.truncated_normal(shape, stddev = 0.01) 
		return tf.Variable(initial)

    #偏置 TensorFlow创建常量tf.constant
	def bias_variable(self, shape):
		initial = tf.constant(0.01, shape = shape)
		return tf.Variable(initial)

	#卷积 tf.nn.conv2d(input, filter, strides, padding, use_cudnn_on_gpu=None, name=None)
	def conv2d(self, x, W):
		return tf.nn.conv2d(x, W, strides = [1, 1, 1, 1], padding = "SAME")

	#池化 tf.nn.max_pool(value, ksize, strides, padding, name=None)
	def max_pool_2x2(self, x):
		return tf.nn.max_pool(x, ksize = [1, 2, 2, 1], strides = [1, 2, 2, 1], padding = "SAME")
	
	#构建CNN模型 inputState QValue
	def createNetwork(self):

		#第一层卷积核
		W_conv1 = self.weight_variable([5, 5, 1, 32])
		b_conv1 = self.bias_variable([32])
		#第二层卷积核
		W_conv2 = self.weight_variable([5, 5, 32, 64])
		b_conv2 = self.bias_variable([64])
		#第三层全连接
		W_fc1 = self.weight_variable([7 * 7 * 64, 1024])
		b_fc1 = self.bias_variable([1024])
		#第四层输出层
		W_fc2 = self.weight_variable([1024, 10])
		b_fc2 = self.bias_variable([10])
		
		# 输入层
		self.inputState = tf.placeholder("float", [None, 784])
		
		x_image = tf.reshape(self.inputState, [-1,28,28,1])
		#第一层卷积层
		h_conv1 = tf.nn.relu(self.conv2d(x_image, W_conv1) + b_conv1)
		h_pool1 = self.max_pool_2x2(h_conv1)
		#第二层卷积层
		h_conv2 = tf.nn.relu(self.conv2d(h_pool1, W_conv2) + b_conv2)
		h_pool2 = self.max_pool_2x2(h_conv2)
		#全连接层
		h_pool2_flat = tf.reshape(h_pool2, [-1, 7*7*64])
		h_fc1 = tf.nn.relu(tf.matmul(h_pool2_flat, W_fc1) + b_fc1)
		
		
		#输出层
		self.y_conv=tf.nn.softmax(tf.matmul(h_fc1, W_fc2) + b_fc2)

		#损失函数以及训练
		self.y_ = tf.placeholder("float", shape=[None, 10])
		self.cross_entropy = -tf.reduce_sum(self.y_*tf.log(self.y_conv))
		self.train_step = tf.train.AdamOptimizer(1e-4).minimize(self.cross_entropy)
		
	def	trainNetwork(self,matrix,label,times):
		print("START Train")
		for i in range(times):
			image = matrix[i]
			labell = label[i]
			self.train_step.run(feed_dict={self.inputState: image, self.y_: labell})
			print("Times: ",i)

		self.saver.save(self.sess,'Model/CNN')
	
	def useCNN(self,image,y):
		y_ = self.y_conv.eval(feed_dict={self.inputState:image})
		y_ = np.asarray(y_)
		y_ = y_.tolist()
		y_index = max(y_)
		y_index = np.argmax(y_index)
		
		labell = np.asarray(y)
		labell = labell.tolist()
		labell_index = max(labell)
		labell_index= np.argmax(labell_index)
		print("预测:" + str(y_index) + " " + "标签" + str(labell_index))
		
		if y_index == labell_index:
			return 1
		else:
			return 0
		
		return 0


def train():
	cnn = CNN()
	matrix = loadImageAndToMatrix(1,4000)
	y = loadYToMatrix(1,4000)
	cnn.trainNetwork(matrix, y,4000)
	print("Finish")


def test():
	cnn = CNN()
	matrix = loadImageAndToMatrix(4001,5000)
	y = loadYToMatrix(4001,5000)
	success = 0
	for i in range(1000):
		image = matrix[i]
		labell = y[i]
		success = success + cnn.useCNN(image, labell)
		
	print("Times:" , success)
	print('Success', success/1000*1.0)


def main():
	#train()
	test()
	

if __name__ == "__main__":
	main()































'''
def loadImage():
	# 读取图片 
	im = Image.open("./Data/1.jpg")

	# 显示图片
	#im.show() 
	
	im = im.convert("1") 
	data = im.getdata()
	data = np.matrix(data)
	print(data.shape)
	#显示二值化后的图片
	new_im = Image.fromarray(np.reshape(data,(28,28)))
	
	new_im.show()
	

#loadImage()
'''





















